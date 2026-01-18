#!/usr/bin/env python3
"""Generate Excel report from existing analysis data."""

from pathlib import Path
import polars as pl
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Paths
PROJECT_ROOT = Path(__file__).parent.parent
ANALYSIS_DIR = PROJECT_ROOT / "analysis"
DATA_DIR = ANALYSIS_DIR / "data"
OUTPUT_FILE = ANALYSIS_DIR / "european_analysis_results.xlsx"


def format_sheet_header(ws, headers):
    """Format header row with styling."""
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
    # Set column widths
    for col_num in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 18


def create_summary_sheet(wb):
    """Create summary sheet with key findings."""
    print("\n  Creating Summary sheet...")
    
    ws = wb.active
    ws.title = "Summary"
    
    # Title
    ws['A1'] = "European Digital Connectivity & Healthcare Accessibility Analysis"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:C1')
    
    row = 3
    ws[f'A{row}'] = "KEY FINDINGS"
    ws[f'A{row}'].font = Font(bold=True, size=14)
    row += 2
    
    # Load data
    conn_df = pl.read_parquet(DATA_DIR / "connectivity_analysis.parquet")
    health_df = pl.read_parquet(DATA_DIR / "healthcare_analysis.parquet")
    demo_df = pl.read_parquet(DATA_DIR / "demographics_analysis.parquet")
    
    # Connectivity summary
    ws[f'A{row}'] = "INTERNET CONNECTIVITY"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    # Europe-wide connectivity
    europe_conn = conn_df.filter(
        (pl.col('region') == 'Europe') &
        (pl.col('metric_type') == 'fixed')
    )
    
    ws[f'A{row}'] = "Metric"
    ws[f'B{row}'] = "Population"
    ws[f'C{row}'] = "Percentage"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'C{row}'].font = Font(bold=True)
    row += 1
    
    # Get each tier
    for tier in ['disconnected', 'very_poor', 'poor', 'basic', 'good']:
        tier_data = europe_conn.filter(pl.col('tier') == tier)
        if len(tier_data) > 0:
            pop = int(tier_data['population'][0])
            pct = float(tier_data['percentage'][0])
            tier_label = tier.replace('_', ' ').title()
            ws[f'A{row}'] = tier_label
            ws[f'B{row}'] = f"{pop:,}"
            ws[f'C{row}'] = f"{pct:.1f}%"
            row += 1
    
    row += 1
    
    # Healthcare summary
    ws[f'A{row}'] = "HEALTHCARE ACCESSIBILITY"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    europe_health = health_df.filter(pl.col('region') == 'Europe')
    
    ws[f'A{row}'] = "Category"
    ws[f'B{row}'] = "Population"
    ws[f'C{row}'] = "Percentage"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'C{row}'].font = Font(bold=True)
    row += 1
    
    for category in ['very_close', 'close', 'moderate', 'far', 'very_far']:
        cat_data = europe_health.filter(pl.col('category') == category)
        if len(cat_data) > 0:
            pop = int(cat_data['population'][0])
            pct = float(cat_data['percentage'][0])
            cat_label = category.replace('_', ' ').title()
            ws[f'A{row}'] = cat_label
            ws[f'B{row}'] = f"{pop:,}"
            ws[f'C{row}'] = f"{pct:.1f}%"
            row += 1
    
    row += 1
    
    # Demographics
    ws[f'A{row}'] = "DEMOGRAPHICS COMPARISON"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    underserved = demo_df.filter(
        (pl.col('region') == 'Europe') & 
        (pl.col('connectivity_group') == 'underserved')
    )
    well_connected = demo_df.filter(
        (pl.col('region') == 'Europe') & 
        (pl.col('connectivity_group') == 'well_connected')
    )
    
    if len(underserved) > 0 and len(well_connected) > 0:
        ws[f'A{row}'] = "Metric"
        ws[f'B{row}'] = "Underserved (<10 Mbps)"
        ws[f'C{row}'] = "Well-Connected (≥100 Mbps)"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'C{row}'].font = Font(bold=True)
        row += 1
        
        under_data = underserved.row(0, named=True)
        well_data = well_connected.row(0, named=True)
        
        ws[f'A{row}'] = "Population"
        ws[f'B{row}'] = f"{int(under_data['total_population']):,}"
        ws[f'C{row}'] = f"{int(well_data['total_population']):,}"
        row += 1
        
        ws[f'A{row}'] = "Population Density (per km²)"
        ws[f'B{row}'] = f"{under_data['pop_density_per_km2']:.1f}" if under_data.get('pop_density_per_km2') else "N/A"
        ws[f'C{row}'] = f"{well_data['pop_density_per_km2']:.1f}" if well_data.get('pop_density_per_km2') else "N/A"
        row += 1
        
        ws[f'A{row}'] = "Mean Healthcare Distance (min)"
        ws[f'B{row}'] = f"{under_data['mean_healthcare_minutes']:.1f}" if under_data.get('mean_healthcare_minutes') else "N/A"
        ws[f'C{row}'] = f"{well_data['mean_healthcare_minutes']:.1f}" if well_data.get('mean_healthcare_minutes') else "N/A"
        row += 1


def create_connectivity_sheets(wb):
    """Create connectivity sheets for each NUTS level."""
    print("\n  Creating Connectivity sheets...")
    
    conn_df = pl.read_parquet(DATA_DIR / "connectivity_analysis.parquet")
    
    # Europe sheet
    print("    Connectivity_Europe...")
    ws = wb.create_sheet("Connectivity_Europe")
    
    headers = ['Metric Type', 'Tier', 'Population', 'Percentage', 'Hexagon Count']
    format_sheet_header(ws, headers)
    
    europe_data = conn_df.filter(pl.col('region') == 'Europe')
    
    row_num = 2
    for row_data in europe_data.iter_rows(named=True):
        ws.cell(row=row_num, column=1, value=row_data['metric_type'])
        ws.cell(row=row_num, column=2, value=row_data['tier'])
        ws.cell(row=row_num, column=3, value=row_data['population'])
        ws.cell(row=row_num, column=4, value=row_data['percentage'])
        ws.cell(row=row_num, column=5, value=row_data['hexagon_count'])
        
        # Format numbers
        ws.cell(row=row_num, column=3).number_format = '#,##0'
        ws.cell(row=row_num, column=4).number_format = '0.0"%"'
        ws.cell(row=row_num, column=5).number_format = '#,##0'
        row_num += 1
    
    # NUTS level sheets
    for level in [0, 1, 2, 3]:
        print(f"    Connectivity_NUTS{level}...")
        ws = wb.create_sheet(f"Connectivity_NUTS{level}")
        
        headers = ['Region', 'Region Name', 'Metric Type', 'Tier', 'Population', 'Percentage', 'Hexagon Count']
        format_sheet_header(ws, headers)
        
        level_data = conn_df.filter(pl.col('nuts_level') == str(level))
        
        row_num = 2
        for row_data in level_data.iter_rows(named=True):
            ws.cell(row=row_num, column=1, value=row_data['region'])
            ws.cell(row=row_num, column=2, value=row_data['region_name'])
            ws.cell(row=row_num, column=3, value=row_data['metric_type'])
            ws.cell(row=row_num, column=4, value=row_data['tier'])
            ws.cell(row=row_num, column=5, value=row_data['population'])
            ws.cell(row=row_num, column=6, value=row_data['percentage'])
            ws.cell(row=row_num, column=7, value=row_data['hexagon_count'])
            
            # Format numbers
            ws.cell(row=row_num, column=5).number_format = '#,##0'
            ws.cell(row=row_num, column=6).number_format = '0.0"%"'
            ws.cell(row=row_num, column=7).number_format = '#,##0'
            row_num += 1


def create_healthcare_sheets(wb):
    """Create healthcare sheets for each NUTS level."""
    print("\n  Creating Healthcare sheets...")
    
    health_df = pl.read_parquet(DATA_DIR / "healthcare_analysis.parquet")
    
    # Europe sheet
    print("    Healthcare_Europe...")
    ws = wb.create_sheet("Healthcare_Europe")
    
    headers = ['Category', 'Population', 'Percentage', 'Hexagon Count']
    format_sheet_header(ws, headers)
    
    europe_data = health_df.filter(pl.col('region') == 'Europe')
    
    row_num = 2
    for row_data in europe_data.iter_rows(named=True):
        ws.cell(row=row_num, column=1, value=row_data['category'])
        ws.cell(row=row_num, column=2, value=row_data['population'])
        ws.cell(row=row_num, column=3, value=row_data['percentage'])
        ws.cell(row=row_num, column=4, value=row_data['hexagon_count'])
        
        # Format numbers
        ws.cell(row=row_num, column=2).number_format = '#,##0'
        ws.cell(row=row_num, column=3).number_format = '0.0"%"'
        ws.cell(row=row_num, column=4).number_format = '#,##0'
        row_num += 1
    
    # NUTS level sheets
    for level in [0, 1, 2, 3]:
        print(f"    Healthcare_NUTS{level}...")
        ws = wb.create_sheet(f"Healthcare_NUTS{level}")
        
        headers = ['Region', 'Region Name', 'Category', 'Population', 'Percentage', 'Hexagon Count']
        format_sheet_header(ws, headers)
        
        level_data = health_df.filter(pl.col('nuts_level') == str(level))
        
        row_num = 2
        for row_data in level_data.iter_rows(named=True):
            ws.cell(row=row_num, column=1, value=row_data['region'])
            ws.cell(row=row_num, column=2, value=row_data['region_name'])
            ws.cell(row=row_num, column=3, value=row_data['category'])
            ws.cell(row=row_num, column=4, value=row_data['population'])
            ws.cell(row=row_num, column=5, value=row_data['percentage'])
            ws.cell(row=row_num, column=6, value=row_data['hexagon_count'])
            
            # Format numbers
            ws.cell(row=row_num, column=4).number_format = '#,##0'
            ws.cell(row=row_num, column=5).number_format = '0.0"%"'
            ws.cell(row=row_num, column=6).number_format = '#,##0'
            row_num += 1


def create_demographics_sheets(wb):
    """Create demographics comparison sheets."""
    print("\n  Creating Demographics sheets...")
    
    demo_df = pl.read_parquet(DATA_DIR / "demographics_analysis.parquet")
    
    ws = wb.create_sheet("Demographics")
    
    headers = ['Region', 'Region Name', 'NUTS Level', 'Connectivity Group', 'Population', 
               'Density (per km²)', 'Mean Healthcare (min)', '% Female', 
               '% Age <15', '% Age 15-64', '% Age 65+']
    format_sheet_header(ws, headers)
    
    row_num = 2
    for row_data in demo_df.iter_rows(named=True):
        ws.cell(row=row_num, column=1, value=row_data['region'])
        ws.cell(row=row_num, column=2, value=row_data['region_name'])
        ws.cell(row=row_num, column=3, value=row_data['nuts_level'])
        ws.cell(row=row_num, column=4, value=row_data['connectivity_group'])
        ws.cell(row=row_num, column=5, value=row_data['total_population'])
        ws.cell(row=row_num, column=6, value=row_data['pop_density_per_km2'])
        ws.cell(row=row_num, column=7, value=row_data['mean_healthcare_minutes'])
        ws.cell(row=row_num, column=8, value=row_data['pct_female'])
        ws.cell(row=row_num, column=9, value=row_data['pct_0_14'])
        ws.cell(row=row_num, column=10, value=row_data['pct_15_64'])
        ws.cell(row=row_num, column=11, value=row_data['pct_65_plus'])
        
        # Format numbers
        ws.cell(row=row_num, column=5).number_format = '#,##0'
        ws.cell(row=row_num, column=6).number_format = '0.0'
        ws.cell(row=row_num, column=7).number_format = '0.0'
        for col in range(8, 12):
            ws.cell(row=row_num, column=col).number_format = '0.0"%"'
        row_num += 1


def create_joint_vulnerability_sheet(wb):
    """Create joint vulnerability analysis sheet."""
    print("\n  Creating Joint_Vulnerability sheet...")
    
    joint_vuln_df = pl.read_parquet(DATA_DIR / "joint_vulnerability.parquet")
    
    ws = wb.create_sheet("Joint_Vulnerability")
    
    headers = ['Region', 'Region Name', 'NUTS Level', 'Vulnerability Type',
               'Threshold (min)', 'Vulnerable Population', '% Vulnerable', 'Total Population']
    format_sheet_header(ws, headers)
    
    row_num = 2
    for row_data in joint_vuln_df.iter_rows(named=True):
        ws.cell(row=row_num, column=1, value=row_data['region'])
        ws.cell(row=row_num, column=2, value=row_data['region_name'])
        ws.cell(row=row_num, column=3, value=row_data['nuts_level'])
        ws.cell(row=row_num, column=4, value=row_data['vulnerability_type'])
        ws.cell(row=row_num, column=5, value=row_data['threshold_minutes'])
        ws.cell(row=row_num, column=6, value=row_data['vulnerable_population'])
        ws.cell(row=row_num, column=7, value=row_data['vulnerable_percentage'])
        ws.cell(row=row_num, column=8, value=row_data['total_population'])
        
        # Format numbers
        ws.cell(row=row_num, column=6).number_format = '#,##0'
        ws.cell(row=row_num, column=7).number_format = '0.0"%"'
        ws.cell(row=row_num, column=8).number_format = '#,##0'
        row_num += 1


def create_excel_report():
    """Create complete Excel report."""
    print("\n" + "="*80)
    print("CREATING EXCEL REPORT")
    print("="*80)
    
    wb = Workbook()
    
    # Create all sheets
    create_summary_sheet(wb)
    create_connectivity_sheets(wb)
    create_healthcare_sheets(wb)
    create_demographics_sheets(wb)
    create_joint_vulnerability_sheet(wb)
    
    # Save
    wb.save(OUTPUT_FILE)
    
    print(f"\n✓ Excel report saved: {OUTPUT_FILE}")
    print(f"  File size: {OUTPUT_FILE.stat().st_size / 1024 / 1024:.1f} MB")
    print(f"  Sheets: {len(wb.sheetnames)}")


if __name__ == "__main__":
    create_excel_report()
    print("\n" + "="*80)
    print("✓ COMPLETED")
    print("="*80)
