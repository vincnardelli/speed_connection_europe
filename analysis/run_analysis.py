#!/usr/bin/env python3
"""
Main orchestrator script for NUTS analysis.

This script runs all analysis steps in sequence and generates
the final Excel file with all results.

Steps:
1. Prepare NUTS data (spatial join)
2. Analyze connectivity
3. Analyze healthcare accessibility
4. Analyze demographics
5. Generate maps and charts
6. Export to single Excel file

Output: analysis/european_analysis_results.xlsx
"""

import sys
from pathlib import Path
import time
import subprocess

import polars as pl
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Paths
PROJECT_ROOT = Path(__file__).parent.parent
ANALYSIS_DIR = PROJECT_ROOT / "analysis"
SCRIPTS_DIR = ANALYSIS_DIR / "scripts"
DATA_DIR = ANALYSIS_DIR / "data"

OUTPUT_FILE = ANALYSIS_DIR / "european_analysis_results.xlsx"

# Analysis scripts
SCRIPTS = [
    "01_prepare_nuts_data.py",
    "02_analyze_connectivity.py",
    "03_analyze_healthcare.py",
    "04_demographics.py",
    "05_generate_maps.py",
]


def run_script(script_name):
    """Run a Python script and capture output."""
    script_path = SCRIPTS_DIR / script_name
    
    print(f"\nRunning {script_name}...")
    print("="*80)
    
    result = subprocess.run(
        [sys.executable, str(script_path)],
        capture_output=False,
        text=True
    )
    
    if result.returncode != 0:
        print(f"✗ {script_name} failed with code {result.returncode}")
        return False
    
    print(f"✓ {script_name} completed successfully")
    return True


def format_sheet_header(ws, headers):
    """Format Excel sheet header row."""
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Auto-width columns
    for col_num in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 15


def create_summary_sheet(wb):
    """Create summary sheet with key findings."""
    print("\n  Creating Summary sheet...")
    
    ws = wb.create_sheet("Summary", 0)
    
    # Load data
    connectivity_df = pl.read_parquet(DATA_DIR / "connectivity_analysis.parquet")
    healthcare_df = pl.read_parquet(DATA_DIR / "healthcare_analysis.parquet")
    demographics_df = pl.read_parquet(DATA_DIR / "demographics_analysis.parquet")
    
    # Europe-level data
    europe_conn = connectivity_df.filter(pl.col('region') == 'Europe')
    europe_health = healthcare_df.filter(pl.col('region') == 'Europe')
    europe_demo = demographics_df.filter(pl.col('region') == 'Europe')
    
    # Write summary
    ws['A1'] = "European NUTS Analysis - Key Findings"
    ws['A1'].font = Font(size=16, bold=True)
    
    row = 3
    
    # Connectivity summary
    ws[f'A{row}'] = "INTERNET CONNECTIVITY"
    ws[f'A{row}'].font = Font(size=14, bold=True, color="366092")
    row += 2
    
    fixed_disconn = europe_conn.filter(
        (pl.col('metric_type') == 'fixed') & 
        (pl.col('tier') == 'disconnected')
    )
    if len(fixed_disconn) > 0:
        pop = int(fixed_disconn['population'][0])
        pct = float(fixed_disconn['percentage'][0])
        ws[f'A{row}'] = "Disconnected (no speed data)"
        ws[f'B{row}'] = f"{pop:,}"
        ws[f'C{row}'] = f"{pct:.1f}%"
        row += 1
    
    fixed_verypoor = europe_conn.filter(
        (pl.col('metric_type') == 'fixed') & 
        (pl.col('tier') == 'very_poor')
    )
    if len(fixed_verypoor) > 0:
        pop = int(fixed_verypoor['population'][0])
        pct = float(fixed_verypoor['percentage'][0])
        ws[f'A{row}'] = "Very Poor (<10 Mbps)"
        ws[f'B{row}'] = f"{pop:,}"
        ws[f'C{row}'] = f"{pct:.1f}%"
        row += 1
    
    # Total underserved
    if len(fixed_disconn) > 0 and len(fixed_verypoor) > 0:
        total_under = int(fixed_disconn['population'][0] + fixed_verypoor['population'][0])
        total_pct = float(fixed_disconn['percentage'][0] + fixed_verypoor['percentage'][0])
        ws[f'A{row}'] = "TOTAL UNDERSERVED"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = f"{total_under:,}"
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'C{row}'] = f"{total_pct:.1f}%"
        ws[f'C{row}'].font = Font(bold=True)
        row += 2
    
    # Healthcare summary
    ws[f'A{row}'] = "HEALTHCARE ACCESSIBILITY"
    ws[f'A{row}'].font = Font(size=14, bold=True, color="366092")
    row += 2
    
    # Load healthcare stats
    healthcare_stats = pl.read_parquet(DATA_DIR / "healthcare_stats.parquet")
    europe_stats = healthcare_stats.filter(pl.col('region') == 'Europe')
    
    for metric in ['mean_distance_minutes', 'pop_gt_5min', 'pop_gt_10min', 'pop_gt_15min', 'pop_gt_30min']:
        metric_data = europe_stats.filter(pl.col('metric_name') == metric)
        if len(metric_data) > 0:
            val = float(metric_data['value'][0]) if metric_data['value'][0] is not None else None
            if val is None:
                continue
                
            if 'distance' in metric:
                ws[f'A{row}'] = "Mean travel time to hospital"
                ws[f'B{row}'] = f"{val:.1f} minutes"
            else:
                threshold = metric.split('_')[-1].replace('min', '')
                ws[f'A{row}'] = f"Population >{threshold} min from hospital"
                ws[f'B{row}'] = f"{int(val):,}"
                
                # Get percentage
                pct_metric = f'{metric}_pct'
                pct_data = europe_stats.filter(pl.col('metric_name') == pct_metric)
                if len(pct_data) > 0 and pct_data['value'][0] is not None:
                    pct_val = float(pct_data['value'][0])
                    ws[f'C{row}'] = f"{pct_val:.1f}%"
            row += 1
    
    row += 1
    
    # Demographics summary
    ws[f'A{row}'] = "DEMOGRAPHICS"
    ws[f'A{row}'].font = Font(size=14, bold=True, color="366092")
    row += 2
    
    underserved_demo = europe_demo.filter(pl.col('connectivity_group') == 'underserved')
    if len(underserved_demo) > 0:
        demo_row = underserved_demo.row(0, named=True)
        total_pop = int(demo_row['total_population']) if demo_row['total_population'] is not None else 0
        pop_density = float(demo_row['pop_density_per_km2']) if demo_row['pop_density_per_km2'] is not None else 0
        pct_65_plus = float(demo_row['pct_65_plus']) if demo_row['pct_65_plus'] is not None else None
        
        ws[f'A{row}'] = "Underserved population"
        ws[f'B{row}'] = f"{total_pop:,}"
        row += 1
        
        ws[f'A{row}'] = "Population density"
        ws[f'B{row}'] = f"{pop_density:.1f} per km²"
        row += 1
        
        if pct_65_plus is not None:
            ws[f'A{row}'] = "Age 65+ (underserved)"
            ws[f'B{row}'] = f"{pct_65_plus:.1f}%"
            row += 1
    
    # Format columns
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15


def create_connectivity_sheets(wb):
    """Create connectivity analysis sheets for all NUTS levels."""
    print("\n  Creating Connectivity sheets...")
    
    connectivity_df = pl.read_parquet(DATA_DIR / "connectivity_analysis.parquet")
    connectivity_stats = pl.read_parquet(DATA_DIR / "connectivity_stats.parquet")
    
    levels = ['All', 0, 1, 2, 3]
    level_names = {
        'All': 'Connectivity_Europe',
        0: 'Connectivity_NUTS0',
        1: 'Connectivity_NUTS1',
        2: 'Connectivity_NUTS2',
        3: 'Connectivity_NUTS3'
    }
    
    for level in levels:
        sheet_name = level_names[level]
        print(f"    {sheet_name}...")
        
        ws = wb.create_sheet(sheet_name)
        
        # Filter data for this level and fixed internet
        level_data = connectivity_df.filter(
            (pl.col('nuts_level') == str(level)) &
            (pl.col('metric_type') == 'fixed')
        )
        
        # Pivot: regions as rows, tiers as columns
        regions = level_data['region'].unique().sort()
        
        # Headers
        headers = ['Region', 'Region Name', 'Disconnected', 'Very Poor (<10 Mbps)', 
                   'Poor (10-25 Mbps)', 'Basic (25-100 Mbps)', 'Good (≥100 Mbps)', 
                   'Total Underserved', 'Mean Speed (Mbps)']
        format_sheet_header(ws, headers)
        
        # Data rows
        row_num = 2
        for region in regions:
            region_data = level_data.filter(pl.col('region') == region)
            
            if len(region_data) == 0:
                continue
            
            region_name = region_data['region_name'][0]
            
            # Get tier populations
            tier_pops = {}
            for tier in ['disconnected', 'very_poor', 'poor', 'basic', 'good']:
                tier_df = region_data.filter(pl.col('tier') == tier)
                tier_pops[tier] = tier_df['population'][0] if len(tier_df) > 0 else 0
            
            # Total underserved
            underserved = tier_pops.get('disconnected', 0) + tier_pops.get('very_poor', 0)
            
            # Mean speed
            stats_data = connectivity_stats.filter(
                (pl.col('region') == region) &
                (pl.col('metric_type') == 'fixed') &
                (pl.col('metric_name') == 'mean_speed_mbps')
            )
            mean_speed = stats_data['value'][0] if len(stats_data) > 0 else None
            
            # Write row
            ws.cell(row=row_num, column=1, value=region)
            ws.cell(row=row_num, column=2, value=region_name)
            ws.cell(row=row_num, column=3, value=tier_pops.get('disconnected', 0))
            ws.cell(row=row_num, column=4, value=tier_pops.get('very_poor', 0))
            ws.cell(row=row_num, column=5, value=tier_pops.get('poor', 0))
            ws.cell(row=row_num, column=6, value=tier_pops.get('basic', 0))
            ws.cell(row=row_num, column=7, value=tier_pops.get('good', 0))
            ws.cell(row=row_num, column=8, value=underserved)
            ws.cell(row=row_num, column=9, value=mean_speed if mean_speed else 'N/A')
            
            # Format numbers
            for col in [3, 4, 5, 6, 7, 8]:
                ws.cell(row=row_num, column=col).number_format = '#,##0'
            if mean_speed:
                ws.cell(row=row_num, column=9).number_format = '0.0'
            
            row_num += 1


def create_healthcare_sheets(wb):
    """Create healthcare analysis sheets for all NUTS levels."""
    print("\n  Creating Healthcare sheets...")
    
    healthcare_df = pl.read_parquet(DATA_DIR / "healthcare_analysis.parquet")
    healthcare_stats = pl.read_parquet(DATA_DIR / "healthcare_stats.parquet")
    
    levels = ['All', 0, 1, 2, 3]
    level_names = {
        'All': 'Healthcare_Europe',
        0: 'Healthcare_NUTS0',
        1: 'Healthcare_NUTS1',
        2: 'Healthcare_NUTS2',
        3: 'Healthcare_NUTS3'
    }
    
    for level in levels:
        sheet_name = level_names[level]
        print(f"    {sheet_name}...")
        
        ws = wb.create_sheet(sheet_name)
        
        # Filter data for this level
        level_data = healthcare_df.filter(pl.col('nuts_level') == str(level))
        
        # Headers
        headers = ['Region', 'Region Name', '<5 min', '5-10 min', '10-15 min', 
                   '15-30 min', '>30 min', 'Mean Distance (min)', '>15 min Total']
        format_sheet_header(ws, headers)
        
        # Data rows
        regions = level_data['region'].unique().sort()
        row_num = 2
        
        for region in regions:
            region_data = level_data.filter(pl.col('region') == region)
            
            if len(region_data) == 0:
                continue
            
            region_name = region_data['region_name'][0]
            
            # Get category populations
            cat_pops = {}
            for cat in ['very_close', 'close', 'moderate', 'far', 'very_far']:
                cat_df = region_data.filter(pl.col('category') == cat)
                cat_pops[cat] = cat_df['population'][0] if len(cat_df) > 0 else 0
            
            # Stats
            stats_data = healthcare_stats.filter(pl.col('region') == region)
            
            mean_dist_data = stats_data.filter(pl.col('metric_name') == 'mean_distance_minutes')
            mean_dist = mean_dist_data['value'][0] if len(mean_dist_data) > 0 else None
            
            gt15_data = stats_data.filter(pl.col('metric_name') == 'pop_gt_15min')
            gt15 = gt15_data['value'][0] if len(gt15_data) > 0 else None
            
            # Write row
            ws.cell(row=row_num, column=1, value=region)
            ws.cell(row=row_num, column=2, value=region_name)
            ws.cell(row=row_num, column=3, value=cat_pops.get('very_close', 0))
            ws.cell(row=row_num, column=4, value=cat_pops.get('close', 0))
            ws.cell(row=row_num, column=5, value=cat_pops.get('moderate', 0))
            ws.cell(row=row_num, column=6, value=cat_pops.get('far', 0))
            ws.cell(row=row_num, column=7, value=cat_pops.get('very_far', 0))
            ws.cell(row=row_num, column=8, value=mean_dist if mean_dist else 'N/A')
            ws.cell(row=row_num, column=9, value=gt15 if gt15 else 'N/A')
            
            # Format numbers
            for col in [3, 4, 5, 6, 7, 9]:
                if ws.cell(row=row_num, column=col).value != 'N/A':
                    ws.cell(row=row_num, column=col).number_format = '#,##0'
            if mean_dist:
                ws.cell(row=row_num, column=8).number_format = '0.0'
            
            row_num += 1


def create_demographics_sheets(wb):
    """Create demographics sheets."""
    print("\n  Creating Demographics sheets...")
    
    demographics_df = pl.read_parquet(DATA_DIR / "demographics_analysis.parquet")
    
    # Europe-level comparison
    europe_data = demographics_df.filter(pl.col('region') == 'Europe')
    
    # Underserved sheet
    ws_under = wb.create_sheet("Demographics_Underserved")
    
    headers = ['Metric', 'Value']
    format_sheet_header(ws_under, headers)
    
    underserved = europe_data.filter(pl.col('connectivity_group') == 'underserved')
    if len(underserved) > 0:
        demo = underserved.row(0, named=True)
        
        metrics = [
            ('Total Population', demo['total_population'], '#,##0'),
            ('Male', demo['pop_male'], '#,##0'),
            ('Female', demo['pop_female'], '#,##0'),
            ('% Female', demo['pct_female'], '0.0"%"'),
            ('Age 0-14', demo['pop_0_14'], '#,##0'),
            ('Age 15-64', demo['pop_15_64'], '#,##0'),
            ('Age 65+', demo['pop_65_plus'], '#,##0'),
            ('% Age 0-14', demo['pct_0_14'], '0.0"%"'),
            ('% Age 15-64', demo['pct_15_64'], '0.0"%"'),
            ('% Age 65+', demo['pct_65_plus'], '0.0"%"'),
            ('Population Density (per km²)', demo['pop_density_per_km2'], '0.0'),
            ('Mean Healthcare Distance (min)', demo['mean_healthcare_minutes'], '0.0'),
            ('Mean Fixed Speed (Mbps)', demo['mean_fixed_speed_mbps'], '0.0'),
        ]
        
        row = 2
        for metric, value, fmt in metrics:
            ws_under.cell(row=row, column=1, value=metric)
            display_val = value if value is not None else 'N/A'
            ws_under.cell(row=row, column=2, value=display_val)
            if value is not None and fmt:
                ws_under.cell(row=row, column=2).number_format = fmt
            row += 1
    
    ws_under.column_dimensions['A'].width = 35
    ws_under.column_dimensions['B'].width = 20
    
    # Well-connected sheet
    ws_conn = wb.create_sheet("Demographics_Connected")
    format_sheet_header(ws_conn, headers)
    
    well_connected = europe_data.filter(pl.col('connectivity_group') == 'well_connected')
    if len(well_connected) > 0:
        demo = well_connected.row(0, named=True)
        
        metrics = [
            ('Total Population', demo['total_population'], '#,##0'),
            ('Male', demo['pop_male'], '#,##0'),
            ('Female', demo['pop_female'], '#,##0'),
            ('% Female', demo['pct_female'], '0.0"%"'),
            ('Age 0-14', demo['pop_0_14'], '#,##0'),
            ('Age 15-64', demo['pop_15_64'], '#,##0'),
            ('Age 65+', demo['pop_65_plus'], '#,##0'),
            ('% Age 0-14', demo['pct_0_14'], '0.0"%"'),
            ('% Age 15-64', demo['pct_15_64'], '0.0"%"'),
            ('% Age 65+', demo['pct_65_plus'], '0.0"%"'),
            ('Population Density (per km²)', demo['pop_density_per_km2'], '0.0'),
            ('Mean Healthcare Distance (min)', demo['mean_healthcare_minutes'], '0.0'),
            ('Mean Fixed Speed (Mbps)', demo['mean_fixed_speed_mbps'], '0.0'),
        ]
        
        row = 2
        for metric, value, fmt in metrics:
            ws_conn.cell(row=row, column=1, value=metric)
            display_val = value if value is not None else 'N/A'
            ws_conn.cell(row=row, column=2, value=display_val)
            if value is not None and fmt:
                ws_conn.cell(row=row, column=2).number_format = fmt
            row += 1
    
    ws_conn.column_dimensions['A'].width = 35
    ws_conn.column_dimensions['B'].width = 20


def create_excel_report():
    """Create comprehensive Excel report with all analysis results."""
    print("\n" + "="*80)
    print("CREATING EXCEL REPORT")
    print("="*80)
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Create sheets
    create_summary_sheet(wb)
    create_connectivity_sheets(wb)
    create_healthcare_sheets(wb)
    create_demographics_sheets(wb)
    
    # Save
    wb.save(OUTPUT_FILE)
    
    file_size = OUTPUT_FILE.stat().st_size / 1024**2
    print(f"\n✓ Excel report saved: {OUTPUT_FILE}")
    print(f"  File size: {file_size:.1f} MB")
    print(f"  Sheets: {len(wb.sheetnames)}")


def main():
    """Main execution."""
    print("="*80)
    print("EUROPEAN NUTS ANALYSIS - COMPLETE PIPELINE")
    print("="*80)
    
    overall_start = time.time()
    
    try:
        # Run all analysis scripts
        for script in SCRIPTS:
            if not run_script(script):
                print(f"\n✗ Pipeline failed at {script}")
                return 1
        
        # Create Excel report
        create_excel_report()
        
        # Summary
        print("\n" + "="*80)
        print("PIPELINE COMPLETED SUCCESSFULLY")
        print("="*80)
        print(f"Total time: {time.time()-overall_start:.1f}s")
        print(f"\nOutputs:")
        print(f"  • Excel report: {OUTPUT_FILE}")
        print(f"  • Maps: {ANALYSIS_DIR / 'figures' / 'maps'}")
        print(f"  • Charts: {ANALYSIS_DIR / 'figures' / 'charts'}")
        print(f"  • Metadata: {ANALYSIS_DIR / 'metadata.md'}")
        print("="*80)
        
        return 0
        
    except Exception as e:
        print(f"\n✗ ERROR: {e}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())
